import requests
import tkinter as tk
from tkinter import simpledialog
from tkinter import ttk 
from tkinter.ttk import Combobox
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
from datetime import datetime
import logging

logger = logging.getLogger("OHLC_Data_gathering")
logger.setLevel(logging.INFO)
console_handler = logging.StreamHandler()
file_handler = logging.FileHandler("OHLC_Data_gathering.log")
console_handler.setLevel(logging.DEBUG)
file_handler.setLevel(logging.INFO)

console_format = logging.Formatter("%(name)s - %(message)s")
file_format = logging.Formatter("%(asctime)s - %(name)s - %(levelname)s - %(message)s")
console_handler.setFormatter(console_format)
file_handler.setFormatter(file_format)

logger.addHandler(console_handler)
logger.addHandler(file_handler)



class AutocompleteCombobox(Combobox):

    def set_completion_list(self, completion_list):
        self._position = 0
        self._hits = []
        self._hits_index = 0
        self._completion_list = sorted(completion_list, key=str.lower)
        self['values'] = self._completion_list
        self.bind('<KeyRelease>', self.handle_keyrelease)
    
    def handle_keyrelease(self, event):
        if event.keysym == 'BackSpace':
            self.delete(self.index(tk.INSERT), tk.END)
            self._position = self.index(tk.END)
            
        if event.keysym == 'Left':
            if self._position < self.index(tk.END):
                self.delete(self._position, tk.END)
            else:
                self._position = self._position - 1
                self.delete(self._position, tk.END)

        if event.keysym == 'Right':
            self._position = self.index(tk.END)
        if len(event.keysym) == 1:
            self.autocomplete()

    def autocomplete(self, delta=0):
        if delta:
            self.delete(self._position, tk.END)
        else:
            self._position = self.index(tk.END)
        _hits = []
        for element in self._completion_list:
            if element.lower().startswith(self.get().lower()):
                _hits.append(element)
        
        if _hits:
            self.delete(0, tk.END)
            self.insert(0, _hits[self._hits_index])
            self.select_range(self._position, tk.END)

class Window(tk.Tk):
    def __init__(self):
        super(Window, self).__init__()
        self.api_key = ""
        self.symbol_name = ""
        self.interval = ""
        self.ohlc = []
        self.use_free_api = tk.IntVar()
        self.use_private_api = tk.IntVar()
        self.all_symbols = []
        self.gathering_all_symbols()
        
        self.time_label = tk.Label(self, text="", bg="blue", fg="white", font=("poppins", 20, "bold"))
        self.time_label.place(x=50, y=25)
        self.show_time()

        free_api = tk.Checkbutton(self, text="Use Free API_KEY", \
                                    command=self.using_free_api, bg="#303133", \
                                        activebackground="#303133", fg='white', \
                                            highlightcolor='black', selectcolor='black', \
                                                variable=self.use_free_api)
        free_api.place(x=50, y=60)

        private_api = tk.Checkbutton(self, text="Use Private API_KEY", \
                                    command=self.using_private_api, bg="#303133", \
                                        activebackground="#303133", fg='white', \
                                            highlightcolor='black', selectcolor='black', \
                                                variable=self.use_private_api)
        private_api.place(x=50, y=90)
        # end api label

        symbol_label = tk.Label(self, text="Sybol Name :", \
                                bg="#303133", fg="#FFFFFF", \
                                font=("poppins", 20, "bold"))
        symbol_label.pack(pady=(10,0))

        # self.symbol = tk.Entry(self, justify="center", \
        #                 font=("poppins", 25, "bold"))
        self.symbol = AutocompleteCombobox(self, font=("poppins", 25, "bold"), width=10, height=15)
        self.symbol.set_completion_list(self.all_symbols)
        self.symbol.place(x=360, y=50)

        search_button = tk.Button(self, text="search", \
                                cursor="hand2", border=12, \
                                    command=self.get_data)
        search_button.place(x=650, y=50)

        tk.Label(self, text="timeframe : ", \
                bg="#303133", fg="#FFFFFF", \
                font=14).place(x=240, y=130)
        self.timeframe = ttk.Spinbox(self, values=["15min", "30min", "1h", "4h"], \
                                command=self.get_timeframe, font=14)
        self.timeframe.place(x=360, y=130)

        tk.Button(self, text="Save in Excel", \
                cursor="hand2", border=12, \
                    command=self.save_excel).place(x=500, y=400)

        tk.Button(self, text="Add Data to existing excel file", \
                cursor="hand2", border=12, \
                    command=self.add_excel).place(x=300, y=400)
        
    def show_time(self):
        text = datetime.now().strftime("%H:%M:%S")
        self.time_label.config(text=text)
        self.after(1000, self.show_time)

    def using_free_api(self):
        self.use_private_api.set(0)
        self.api_key = "demo"
        logger.info("Using free api")

    def using_private_api(self):
        self.use_free_api.set(0)
        self.api_key = simpledialog.askstring("set api key", "Enter your API key")
        logger.info("Using private api with api key : %s", self.api_key)

    def get_timeframe(self):
        interval = self.timeframe.get()
        return interval
    
    def gathering_all_symbols(self):
        logger.info("Loading Symbols")
        categories = ["forex_pairs", "commodities", "cryptocurrencies", "indices"]
        for category in categories:
            url = f"https://api.twelvedata.com/{category}"
            response = requests.get(url)
            data = response.json()
            for symbol in data['data']:
                self.all_symbols.append(symbol['symbol'])

    def get_data(self):
        self.symbol_name = self.symbol.get()
        self.interval = self.get_timeframe()
        if messagebox.askokcancel("Config", f"symbol : '{self.symbol_name}'\ntimeframe : '{self.interval}'"):
            try:
                url = f"https://api.twelvedata.com/quote?symbol={self.symbol_name}&interval={self.interval}&apikey={self.api_key}"
                response = requests.get(url)
                data = response.json()
                logger.info("Data arrived")
            except Exception as e:
                messagebox.showerror("Error", "Something Happend in getting data")
                logger.error(f"Error in getting data : {e}")
        else:
            return
        try:
            logger.info("inserting data")
            time = datetime.fromtimestamp(data['timestamp']).strftime("%H:%M:%S")
            self.ohlc = [time, data['open'], data['high'], data['low'], data['close']]
            tk.Label(window, text="Time: ", bg="#303133", fg="#FFFFFF", font=("poppins", 12, "bold")).place(x=300, y=220)
            tk.Label(window, text=self.ohlc[0], bg="#303133", fg="#FFFFFF", font=("poppins", 12, "bold")).place(x=350, y=220)
            tk.Label(window, text="Open: ", bg="#303133", fg="#FFFFFF", font=("poppins", 12, "bold")).place(x=100, y=300)
            tk.Label(window, text=self.ohlc[1], bg="#303133", fg="#FFFFFF", font=("poppins", 12, "bold")).place(x=150, y=300)
            tk.Label(window, text="High: ", bg="#303133", fg="#FFFFFF", font=("poppins", 12, "bold")).place(x=300, y=300)
            tk.Label(window, text=self.ohlc[2], bg="#303133", fg="#FFFFFF", font=("poppins", 12, "bold")).place(x=350, y=300)
            tk.Label(window, text="Low: ", bg="#303133", fg="#FFFFFF", font=("poppins", 12, "bold")).place(x=500, y=300)
            tk.Label(window, text=self.ohlc[3], bg="#303133", fg="#FFFFFF", font=("poppins", 12, "bold")).place(x=550, y=300)
            tk.Label(window, text="Close: ", bg="#303133", fg="#FFFFFF", font=("poppins", 12, "bold")).place(x=700, y=300)
            tk.Label(window, text=self.ohlc[4], bg="#303133", fg="#FFFFFF", font=("poppins", 12, "bold")).place(x=750, y=300)
        except Exception as e:
            messagebox.showwarning("Data Not Recieved", "Change API_KEY")
            logger.error(f"Data Not Recieved : {e}")

    def save_excel(self):
        logger.info("Saving data")
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "time"
        ws["B1"] = "open"
        ws["C1"] = "high"
        ws["D1"] = "low"
        ws["E1"] = "close"
        
        for row in ws.iter_rows(min_row=2, max_row=2, max_col=5):
            i = 0
            for cell in row:
                cell.value = self.ohlc[i]
                i += 1

        for col in ws.iter_cols(min_col=6, max_col=6, min_row=2, max_row=2):
            for cell in col:
                cell.value = self.interval
        
        if '/' in self.symbol_name:
            file_name = self.symbol_name.replace('/', '_')
        else:
            file_name = self.symbol_name

        wb.save(f"{file_name}.xlsx")
        messagebox.showinfo("completed", "All Done!")
        logger.info("All Done!")
        
    def add_excel(self):
        logger.info("Adding data")    
        if '/' in self.symbol_name:
            file_name = self.symbol_name.replace('/', '_')
        else:
            file_name = self.symbol_name

        wb = load_workbook(f"{file_name}.xlsx")
        ws = wb.active

        ohlc2 = self.ohlc.copy()
        ohlc2.append(self.interval)
        ohlc3 = [ohlc2]
        for _ in ohlc3:
            ws.append(_)

        wb.save(f"{file_name}.xlsx")
        messagebox.showinfo("Completed", "Data Added to Excel File.")
        logger.info("Data Added to Excel File")

if __name__ == "__main__":
    window = Window()
    window.title("OHLC Data")
    window.config(bg="#303133")
    window.geometry("900x500+300+150")
    window.bind("<Control-q>", lambda event=None:window.destroy())
    window.bind("<Control-Q>", lambda event=None:window.destroy())
    window.mainloop()